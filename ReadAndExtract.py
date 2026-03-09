import os
import PyPDF2
import re
import math
from datetime import datetime
# openpyxl must be installed for Excel export (pip install openpyxl)
import openpyxl
from openpyxl.utils import get_column_letter

# Regex for a valid transaction line that could be a purchase.
transaction_pattern = re.compile(
    r"^(\d{2}-[A-Za-z]{3}-\d{4})\s+"  # Date (e.g., 18-Jan-2021)
    r"(\(?[0-9,.-]+\)?|--)\s+"        # Amount (can be parenthesized for negative)
    r"(.+?)\s+"                      # Price, Units, and Description (non-greedy)
    r"([0-9,.-]+)$"                  # Unit Balance
)

# A simpler pattern for stamp duty to ensure it's caught
stamp_duty_pattern = re.compile(
    r"^\d{2}-[A-Za-z]{3}-\d{4}\s+([\d,.-]+)\s+\*\*\* Stamp Duty \*\*\*"
)

# Regex for lines that should be ignored during transaction parsing
# This is now more comprehensive to catch all parts of a page header.
ignore_pattern = re.compile(
    r"^(Date\s+Amount|CAMSCASWS-.*|Total|Phone Res:|Mobile:|Email Id:|Page\s+\d+|^\(INR\)|\d{2}-[A-Za-z]{3}-\d{4}\s+To\s+\d{2}-[A-Za-z]{3}-\d{4}|To\s\d{2,}|PAN:)"
)

# Regex to detect the end of a scheme's transaction list
end_of_scheme_pattern = re.compile(r"^(NAV on|Closing Unit Balance:|W\.e\.f:|Exit Load:)")

class MFTransaction:
    """Represents a simplified purchase transaction with date and combined amount."""
    def __init__(self, date, amount):
        self.date: datetime = date
        self.amount: float = amount

    def __repr__(self):
        return f"Transaction({self.date.strftime('%Y-%m-%d')}, Amount: {self.amount})"

class MFScheme:
    """Represents a single mutual fund scheme, containing its details and transactions."""
    def __init__(self, name, folio):
        self.name: str = name
        self.folio: str = folio
        self.transactions: list[MFTransaction] = []
        self.current_market_value: float = 0.0

    def __repr__(self):
        return (f"Scheme('{self.name}', Folio: '{self.folio}', "
                f"MarketValue: {self.current_market_value}, Transactions: {len(self.transactions)})")

def parse_float(s: str) -> float:
    """Converts a string from the PDF to a float, handling commas and parentheses for negatives."""
    s = s.strip().replace(",", "")
    if s == '***' or s == '--': return 0.0
    if s.startswith("(") and s.endswith(")"): return -float(s[1:-1])
    try: return float(s)
    except ValueError: return 0.0

def sanitize_sheet_name(name):
    """Sanitizes a string to be a valid Excel sheet name."""
    invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
    for char in invalid_chars:
        name = name.replace(char, '')
    return name[:31]

def save_to_excel(mf_data, filename):
    """Saves the extracted MF data into an Excel file with custom formatting for each scheme."""
    workbook = openpyxl.Workbook()
    if "Sheet" in workbook.sheetnames: workbook.remove(workbook["Sheet"])

    for mf_house, schemes in mf_data.items():
        sheet_name = sanitize_sheet_name(mf_house)
        sheet = workbook.create_sheet(title=sheet_name)
        for i, scheme in enumerate(schemes):
            if i > 0: sheet.append([]); sheet.append(["*" * 80]); sheet.append([])
            sheet.append(["Scheme Name", scheme.name])
            sheet.append(["Folio", scheme.folio])
            sheet.append(["Current Market Value", scheme.current_market_value])
            sheet.append([])
            if scheme.transactions:
                sheet.append(["Transaction Date", "Transaction Amount"])
                for transaction in scheme.transactions:
                    sheet.append([transaction.date.strftime("%d-%m-%Y"), transaction.amount])
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 20
    try:
        workbook.save(filename)
        print(f"Successfully saved data to {filename}")
    except Exception as e:
        print(f"Error saving Excel file {filename}: {e}")

def Extract_data_from_the_file(file_path, pdf_reader):
    """
    Extracts and structures transaction data from the provided PDF reader object.
    """
    print(f"Processing the file: {os.path.basename(file_path)}")
    
    MfData: dict[str, list[MFScheme]] = {}
    full_text = ""
    for page in pdf_reader.pages:
        full_text += page.extract_text() + "\n"

    in_summary_section = False
    mf_summary_list = []
    for line in full_text.split('\n'):
        if "PORTFOLIO SUMMARY" in line: in_summary_section = True; continue
        if "Total" in line and in_summary_section:
            in_summary_section = False
            summary_end_pos = full_text.find(line) + len(line)
            break
        if in_summary_section:
            match = re.match(r"\s*([^0-9]+?Mutual Fund)\s", line, re.IGNORECASE)
            if match:
                mf_name = match.group(1).strip()
                if mf_name not in mf_summary_list: mf_summary_list.append(mf_name)
    
    mf_section_starts = []
    for mf_house in mf_summary_list:
        try:
            for match in re.finditer(re.escape(mf_house), full_text[summary_end_pos:], re.IGNORECASE):
                actual_start = summary_end_pos + match.start()
                actual_end = summary_end_pos + match.end()
                lookahead_text = full_text[actual_end:actual_end + 400]
                if "PAN:" in lookahead_text and "KYC:" in lookahead_text:
                    mf_section_starts.append((mf_house, actual_start)); break
        except re.error: pass
            
    mf_section_starts.sort(key=lambda x: x[1])

    for i, (mf_house, start_index) in enumerate(mf_section_starts):
        end_index = mf_section_starts[i+1][1] if i + 1 < len(mf_section_starts) else len(full_text)
        mf_block_text = full_text[start_index:end_index]
        MfData[mf_house] = []
        scheme_pattern = re.compile(r"([^\n]+(?:[ \t]*\n[^\n]+)?)\s*- ISIN:[\s\S]*?Folio No:\s*([\w\s/]+)")
        scheme_matches = list(scheme_pattern.finditer(mf_block_text))

        for j, scheme_match in enumerate(scheme_matches):
            raw_name = scheme_match.group(1).strip()
            potential_name = raw_name.split('\n')[-1].strip()
            name_parts = potential_name.split('-', 1)
            if len(name_parts) > 1 and len(name_parts[0].strip()) < 15 and ' ' not in name_parts[0].strip():
                scheme_name = name_parts[1].strip()
            else: scheme_name = potential_name
            folio = scheme_match.group(2).strip().split('\n')[0].split('/')[0].strip()
            scheme_obj = MFScheme(name=scheme_name, folio=folio)
            MfData[mf_house].append(scheme_obj)

            trans_start_index = scheme_match.end()
            next_scheme_start = len(mf_block_text)
            if j + 1 < len(scheme_matches): next_scheme_start = scheme_matches[j+1].start()
            transactions_text = mf_block_text[trans_start_index:next_scheme_start]
            market_value_match = re.search(r"Market Value on .*?:\s*INR\s*([\d,.-]+)", transactions_text)
            if market_value_match: scheme_obj.current_market_value = math.floor(parse_float(market_value_match.group(1)))
            lines = transactions_text.split('\n')
            k = 0
            while k < len(lines):
                line = lines[k].strip()
                # Check for ignore pattern more robustly
                if not line or end_of_scheme_pattern.search(line):
                    k += 1; continue
                if ignore_pattern.search(line): 
                    k += 1; continue
                
                trans_match = transaction_pattern.match(line)
                if not trans_match: k+= 1; continue
                try:
                    date_str, amount_str, rest_of_line, balance_str = trans_match.groups()
                    desc_match = re.search(r"([A-Za-z].*)", rest_of_line)
                    if not desc_match: k+=1; continue
                    desc = desc_match.group(1).strip()

                    if "Purchase" in desc or "Systematic Investment" in desc:
                        purchase_amount = parse_float(amount_str)
                        next_k = k + 1
                        while next_k < len(lines):
                            next_line = lines[next_k].strip()
                            # Use the same robust ignore logic here
                            if not next_line or end_of_scheme_pattern.search(next_line):
                                next_k += 1; continue
                            if ignore_pattern.search(next_line):
                                next_k += 1; continue
                            
                            if "*** Stamp Duty ***" in next_line:
                                # Simple split to get amount is more robust
                                parts = next_line.split()
                                if len(parts) > 1:
                                    try:
                                        stamp_duty_amount = parse_float(parts[1])
                                        purchase_amount += stamp_duty_amount
                                        k = next_k # Consume stamp duty line
                                    except (ValueError, IndexError): pass
                            
                            # Found the next meaningful line, so stop scanning
                            break
                        
                        scheme_obj.transactions.append(MFTransaction(datetime.strptime(date_str, "%d-%b-%Y"), purchase_amount))
                except (ValueError, IndexError): pass
                k += 1

    return MfData

def main():
    """
    Reads all PDF files in the 'Resources' directory, using a password
    from 'password.txt' to open them, and then passes them to a processing function.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    resources_path = os.path.join(script_dir, "Resources")
    password_file_path = os.path.join(resources_path, "password.txt")
    
    try:
        with open(password_file_path, 'r') as f: password = f.read().strip()
    except FileNotFoundError: print(f"Error: Password file not found at {password_file_path}"); return
    try:
        files = os.listdir(resources_path)
    except FileNotFoundError: print(f"Error: Resources directory not found at {resources_path}"); return

    for file_name in files:
        if file_name.lower().endswith(".pdf"):
            file_path = os.path.join(resources_path, file_name)
            try:
                with open(file_path, 'rb') as pdf_file:
                    pdf_reader = PyPDF2.PdfReader(pdf_file)
                    mfdata_from_file = None
                    if pdf_reader.is_encrypted:
                        if not pdf_reader.decrypt(password):
                            print(f"Wrong password for file: {file_name}"); continue
                    
                    mfdata_from_file = Extract_data_from_the_file(file_path, pdf_reader)
                    
                    if mfdata_from_file:
                        excel_filename = os.path.join(resources_path, file_name.replace('.pdf', '.xlsx'))
                        save_to_excel(mfdata_from_file, excel_filename)

            except Exception as e:
                print(f"Could not process file {file_name}. Error: {e}")

if __name__ == "__main__":
    main()
